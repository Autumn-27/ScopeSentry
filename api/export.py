# -------------------------------------
# @file      : export.py
# @author    : Autumn
# @contact   : rainy-autumn@outlook.com
# @time      : 2024/6/16 16:11
# -------------------------------------------
import base64
import json
import os
import traceback

import aiofiles
from bson import ObjectId
from fastapi import APIRouter, Depends, BackgroundTasks
from openpyxl.utils.exceptions import IllegalCharacterError
from starlette.responses import FileResponse

from api.users import verify_token
from motor.motor_asyncio import AsyncIOMotorCursor
from core.db import get_mongo_db, get_project
import pandas as pd

from core.default import FIELD
from core.util import *
from pymongo import ASCENDING, DESCENDING, results
from loguru import logger
from openpyxl import Workbook

router = APIRouter()


@router.post("/export")
async def export_data(request_data: dict, db=Depends(get_mongo_db), _: dict = Depends(verify_token),
                      background_tasks: BackgroundTasks = BackgroundTasks()):
    index = request_data.get("index", "")
    quantity = int(request_data.get("quantity", 0))
    export_type = request_data.get("type", "")
    search_query = request_data.get("search", "")
    if index == "" or quantity == 0 or export_type == "":
        return {"code": 500, "message": f"get index, quantity, export_type null"}
    query = {}
    if export_type == "search":
        query = await get_search_query(index, request_data)
        if query == "" or query is None:
            return {"message": "Search condition parsing error", "code": 500}
    if index == "PageMonitoring":
        query["diff"] = {"$ne": []}
    file_name = generate_random_string(16)
    file_type = request_data.get("filetype", "csv")
    if file_type != "csv" and file_type != "json":
        return {"message": "filetype error", "code": 500}
    fields = request_data.get("field", [])
    if len(fields) == 0:
        return {"message": "fields is null", "code": 500}
    result = await db.export.insert_one({
        "file_name": file_name + "." + file_type,
        "create_time": get_now_time(),
        "quantity": quantity,
        "data_type": index,
        "file_type": file_type,
        "state": 0,
        "end_time": "",
        "file_size": ""
    })
    if result.inserted_id:
        background_tasks.add_task(export_data_from_mongodb2, quantity, query, file_name, index, file_type, fields)
        return {"message": "Successfully added data export task", "code": 200}
    else:
        return {"message": "Failed to export data", "code": 500}


@router.post("/getfield")
async def export_data(request_data: dict, db=Depends(get_mongo_db), _: dict = Depends(verify_token),
                      background_tasks: BackgroundTasks = BackgroundTasks()):
    index = request_data["index"]
    if index in FIELD:
        return {"code": 200, "data": {"field": FIELD[index]}}


async def fetch_data(db, collection, query, quantity, fields):
    # 预先构造替换映射
    # project_map = {original_value: new_value for new_value, original_value in project_list.items()}
    fields_dict = {}
    if collection == "asset":
        fields_dict["type"] = 1
    for i in fields:
        fields_dict[i] = 1
    # 使用 $cond 和 $in 来避免复杂的数组索引操作
    pipeline = [
        {"$match": query},
        {"$limit": quantity},
        # {"$addFields": {
        #     "project": {
        #         "$cond": {
        #             "if": {"$in": ["$project", list(project_map.keys())]},  # 检查 project 是否在 project_map 中
        #             "then": {
        #                 "$let": {
        #                     "vars": {
        #                         "mapped_project": {
        #                             "$arrayElemAt": [
        #                                 list(project_map.values()),
        #                                 {"$indexOfArray": [list(project_map.keys()), "$project"]}
        #                             ]
        #                         }
        #                     },
        #                     "in": "$$mapped_project"  # 如果匹配到，则替换为对应的新值
        #                 }
        #             },
        #             "else": "$project"  # 如果没有匹配到，则保留原值
        #         }
        #     }
        # }},
        {"$project": fields_dict}
    ]

    cursor = db[collection].aggregate(pipeline)
    return cursor


def flatten_dict(d):
    items = []
    for k, v in d.items():
        if isinstance(v, dict):
            items.append((k, str(v)))
        elif isinstance(v, list):
            items.append((k, ', '.join(map(str, v))))
        else:
            items.append((k, v))
    return dict(items)


def clean_string(value):
    if isinstance(value, str):
        # 过滤掉非法字符（ASCII码 < 32 或 >= 127），增加支持中文显示
        return ''.join(char for char in value if (32 <= ord(char) < 127) or (0x4E00 <= ord(char) <= 0x9FFF))
    return value


async def export_data_from_mongodb2(quantity, query, file_name, index, file_type, fields):
    logger.info("导出开始")
    async for db in get_mongo_db():
        try:
            cursor = await fetch_data(db, index, query, quantity, fields)
            relative_path = f'file/{file_name}.{file_type}'
            file_path = os.path.join(os.getcwd(), relative_path)
            if file_type == "csv":
                await export_to_excel(cursor, file_path, index, fields)
            elif file_type == "json":
                await export_to_json(cursor, file_path)
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # kb
            update_document = {
                "$set": {
                    "state": 1,
                    "end_time": get_now_time(),
                    "file_size": str(round(file_size, 2))
                }
            }
            await db.export.update_one({"file_name": file_name + "." + file_type}, update_document)
        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            update_document = {
                "$set": {
                    "state": 2,
                }
            }
            await db.export.update_one({"file_name": file_name + "." + file_type}, update_document)
    logger.info("导出结束")


def convert_bytes_to_base64(data):
    """将 bytes 类型转换为 base64 编码的字符串"""
    if isinstance(data, bytes):
        return base64.b64encode(data).decode('utf-8')  # 返回 base64 编码的字符串
    return data


def convert_datetime_to_str(data):
    """将 datetime 类型转换为字符串"""
    if isinstance(data, datetime):
        return data.isoformat()  # 使用 ISO 8601 格式（例如：2025-02-18T10:30:00）
    return data


def process_document(document):
    """递归处理文档中的所有 bytes 和 datetime 字段"""
    if isinstance(document, dict):
        return {key: process_document(value) for key, value in document.items()}
    elif isinstance(document, list):
        return [process_document(item) for item in document]
    elif isinstance(document, bytes):
        return convert_bytes_to_base64(document)
    elif isinstance(document, datetime):
        return convert_datetime_to_str(document)
    return document


async def export_to_json(cursor, file_path):
    # 使用异步文件操作
    async with aiofiles.open(file_path, 'a', encoding='gbk', errors='ignore') as file:
        async for document in cursor:
            # 将 ObjectId 转换为字符串
            document["_id"] = str(document["_id"])
            # 递归处理 document 中的 bytes 和 datetime 字段
            document = process_document(document)
            # 写入 JSON 文件
            await file.write(json.dumps(document, ensure_ascii=False) + "\n")


async def export_to_excel(cursor, file_path, index, fields):
    wb = Workbook()
    if index == "asset":
        http_header = []
        other_header = []
        for f in fields:
            if f in FIELD["httpAsset"]:
                http_header.append(f)
            if f in FIELD["otherAsset"]:
                other_header.append(f)
        # 创建两个工作表
        http_ws = wb.active
        http_ws.title = 'HTTP Data'
        other_ws = wb.create_sheet(title='Other Data')

        # 写入HTTP Data列名
        http_ws.append(http_header)
        # 写入Other Data列名
        other_ws.append(other_header)
        async for document in cursor:
            flattened_doc = flatten_dict(document)
            if flattened_doc["type"] == "http":
                row = [clean_string(flattened_doc.get(col, "")) for col in http_header]
                http_ws.append(row)
            else:
                row = [clean_string(flattened_doc.get(col, "")) for col in other_header]
                other_ws.append(row)
    else:
        header = []
        for f in fields:
            if f in FIELD[index]:
                header.append(f)
        ws = wb.active
        ws.title = index
        ws.append(header)
        async for document in cursor:
            flattened_doc = flatten_dict(document)
            row = [clean_string(flattened_doc.get(col, "")) for col in header]
            ws.append(row)
    try:
        wb.save(file_path)
        logger.info(f"Data saved to {file_path} successfully.")
    except IllegalCharacterError as e:
        logger.error("导出内容有不可见字符，忽略此错误")


async def export_data_from_mongodb(quantity, query, file_name, index):
    logger.info("导出开始")
    async for db in get_mongo_db():
        try:
            # global Project_List
            # if len(Project_List) == 0:
            #     await get_project(db)
            cursor = await fetch_data(db, index, query, quantity)
            result = await cursor.to_list(length=None)
            relative_path = f'file/{file_name}.xlsx'
            file_path = os.path.join(os.getcwd(), relative_path)
            wb = Workbook()
            if index == "asset":
                http_columns = {
                    "time": "时间",
                    "lastScanTime": "最近扫描时间",
                    "tls": "TLS_Data",
                    "hash": "Hash",
                    "cdnname": "Cdn_Name",
                    "port": "端口",
                    "url": "url",
                    "title": "标题",
                    "type": "类型",
                    "error": "错误",
                    "body": "响应体",
                    "host": "域名",
                    "ip": "IP",
                    "screenshot": "截图",
                    "faviconmmh3": "图标Hash",
                    "faviconpath": "faviconpath",
                    "rawheaders": "响应头",
                    "jarm": "jarm",
                    "technologies": "technologies",
                    "statuscode": "响应码",
                    "contentlength": "contentlength",
                    "cdn": "cdn",
                    "webcheck": "webcheck",
                    "project": "项目",
                    "iconcontent": "图标",
                    "taskName": "任务",
                    "webServer": "webServer",
                    "service": "service",
                    "rootDomain": "根域名",
                    "tags": "tags"
                }
                other_columns = {
                    "time": "时间",
                    "lastScanTime": "最近扫描时间",
                    "host": "域名",
                    "ip": "IP",
                    "port": "端口",
                    "service": "服务",
                    "tls": "TLS",
                    "transport": "transport",
                    "version": "版本",
                    "metadata": "metadata",
                    "project": "项目",
                    "type": "类型",
                    "tags": "tags",
                    "taskName": "任务",
                    "rootDomain": "根域名",
                }
                # 创建两个工作表
                http_ws = wb.active
                http_ws.title = 'HTTP Data'
                other_ws = wb.create_sheet(title='Other Data')

                # 写入HTTP Data列名
                http_ws.append(list(http_columns.values()))
                # 写入Other Data列名
                other_ws.append(list(other_columns.values()))

                # 分别写入数据
                for doc in result:
                    flattened_doc = flatten_dict(doc)
                    if doc["type"] == "other":
                        row = [clean_string(flattened_doc.get(col, "")) for col in other_columns.keys()]
                        other_ws.append(row)
                    else:
                        row = [clean_string(flattened_doc.get(col, "")) for col in http_columns.keys()]
                        http_ws.append(row)
            else:
                columns = {}
                if index == "subdomain":
                    columns = {'host': '域名', 'type': '解析类型', 'value': '解析值', 'ip': '解析IP', 'project': '项目',
                               'time': '时间', "taskName": "任务", "tags": "tags", "rootDomain": "根域名"}
                if index == "SubdoaminTakerResult":
                    columns = {
                        'input': '源域名', 'value': '解析值', 'cname': '接管类型', 'response': '响应体',
                        'project': '项目', "taskName": "任务", "tags": "tags", "rootDomain": "根域名"
                    }
                if index == "UrlScan":
                    columns = {
                        'input': '输入', 'source': '来源', 'outputtype': '输出类型', 'output': '输出',
                        'statuscode': 'statuscode', 'length': 'length', 'time': '时间', 'project': '项目',
                        "taskName": "任务", "tags": "tags", "rootDomain": "根域名"
                    }
                if index == "crawler":
                    columns = {
                        'url': 'URL', 'method': 'Method', 'body': 'Body', 'project': '项目', "taskName": "任务",
                        "tags": "tags", "rootDomain": "根域名"
                    }
                if index == "SensitiveResult":
                    columns = {
                        'url': 'URL', 'sid': '规则名称', 'match': '匹配内容', 'project': '项目', 'body': '响应体',
                        'color': '等级', 'time': '时间', 'md5': '响应体MD5', "taskName": "任务", "tags": "tags",
                        "rootDomain": "根域名"
                    }
                if index == "DirScanResult":
                    columns = {
                        'url': 'URL', 'status': '响应码', 'msg': '跳转', 'project': '项目', "taskName": "任务",
                        "tags": "tags", "rootDomain": "根域名"
                    }
                if index == "vulnerability":
                    columns = {
                        'url': 'URL', 'vulname': '漏洞', 'matched': '匹配', 'project': '项目', 'level': '危害等级',
                        'time': '时间', 'request': '请求', 'response': '响应', "taskName": "任务", "tags": "tags",
                        "rootDomain": "根域名"
                    }
                if index == "PageMonitoring":
                    columns = {
                        'url': 'URL', 'content': '响应体', 'hash': '响应体Hash', 'diff': 'Diff',
                        'state': '状态', 'project': '项目', 'time': '时间', "taskName": "任务", "tags": "tags",
                        "rootDomain": "根域名"
                    }
                ws = wb.active
                ws.title = index
                ws.append(list(columns.values()))

                for doc in result:
                    flattened_doc = flatten_dict(doc)
                    row = [clean_string(flattened_doc.get(col, "")) for col in columns.keys()]
                    ws.append(row)
            try:
                wb.save(file_path)
                logger.info(f"Data saved to {file_path} successfully.")
            except IllegalCharacterError as e:
                logger.error("导出内容有不可见字符，忽略此错误")
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # kb
            update_document = {
                "$set": {
                    "state": 1,
                    "end_time": get_now_time(),
                    "file_size": str(round(file_size, 2))
                }
            }
            await db.export.update_one({"file_name": file_name}, update_document)
        except Exception as e:
            logger.error(str(e))
            logger.error(traceback.format_exc())
            update_document = {
                "$set": {
                    "state": 2,
                }
            }
            await db.export.update_one({"file_name": file_name}, update_document)
    logger.info("导出结束")


@router.get("/export/record")
async def get_export_record(db=Depends(get_mongo_db), _: dict = Depends(verify_token)):
    cursor: AsyncIOMotorCursor = db.export.find({},
                                                {"_id": 0, "id": {"$toString": "$_id"}, "file_name": 1, "end_time": 1,
                                                 "create_time": 1, "data_type": 1, "state": 1, 'file_size': 1, "file_type": 1}).sort(
        [("create_time", DESCENDING)])
    result = await cursor.to_list(length=None)
    return {
        "code": 200,
        "data": {
            'list': result
        }
    }


@router.post("/export/delete")
async def delete_export(request_data: dict, db=Depends(get_mongo_db), _: dict = Depends(verify_token)):
    try:
        export_ids = request_data.get("ids", [])
        delete_filename = []
        for id in export_ids:
            if is_valid_filename(id):
                relative_path = f'file/{id}'
                file_path = os.path.join(os.getcwd(), relative_path)
                if os.path.exists(file_path):
                    os.remove(file_path)
                delete_filename.append(id)
        if len(delete_filename) == 0:
            return {"code": 404, "message": "Export file not found"}
        result = await db.export.delete_many({"file_name": {"$in": delete_filename}})

        if result.deleted_count > 0:
            return {"code": 200, "message": "Export file deleted successfully"}
        else:
            return {"code": 404, "message": "Export file not found"}

    except Exception as e:
        logger.error(str(e))
        # Handle exceptions as needed
        return {"message": "error", "code": 500}


def is_valid_filename(name):
    try:
        tmp= name.split(".")
        file_name = tmp[0]
        file_type = tmp[1]
        if file_type != "json" and file_type != "csv":
            return False
        return is_valid_string(file_name)
    except:
        return False


@router.get("/export/download")
async def download_export(file_name: str):
    if is_valid_filename(file_name):
        relative_path = f'file/{file_name}'
        file_path = os.path.join(os.getcwd(), relative_path)
        if os.path.exists(file_path):
            return FileResponse(path=file_path, filename=file_name)
        else:
            return {"message": "file not found", "code": 500}
    else:
        return {"message": "file not found", "code": 500}
